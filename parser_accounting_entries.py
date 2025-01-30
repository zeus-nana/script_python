import json
from openpyxl import load_workbook

def is_numeric_account(value):
    if not value:
        return False
    return str(value).replace('0', '').isdigit()

def get_formula_value(ws, cell):
    if isinstance(cell.value, str) and cell.value.startswith('='):
        formula_ref = cell.value[1:]
        try:
            referenced_cell = ws[formula_ref]
            return referenced_cell.value
        except:
            return cell.value
    return cell.value

def process_excel_to_json(file_path):
    wb = load_workbook(file_path, data_only=True)
    result = {}

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        entries = []
        current_entry = None

        def get_merged_cell_value(cell):
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    return ws.cell(merged_range.min_row, merged_range.min_col).value
            return cell.value

        for row in range(1, ws.max_row + 1):
            compte_debit = ws.cell(row, 2).value
            compte_credit = ws.cell(row, 3).value

            if not (is_numeric_account(compte_debit) or is_numeric_account(compte_credit)):
                continue

            libelle = get_merged_cell_value(ws.cell(row, 4))
            debit_info = get_formula_value(ws, ws.cell(row, 5))
            credit_info = get_formula_value(ws, ws.cell(row, 6))

            if compte_debit or compte_credit:
                if current_entry and current_entry["comptes"]:
                    entries.append(current_entry)

                merged_libelle = None
                for i in range(row + 1, min(row + 5, ws.max_row + 1)):
                    cell = ws.cell(i, 4)
                    potential_libelle = get_merged_cell_value(cell)
                    if potential_libelle and isinstance(potential_libelle, str) and "Dépôt" in potential_libelle:
                        merged_libelle = potential_libelle
                        break

                current_entry = {
                    "ligne": row,
                    "comptes": [],
                    "libelle": merged_libelle
                }

                if compte_debit:
                    current_entry["comptes"].append({
                        "numero": str(compte_debit),
                        "sens": "debit",
                        "intitule": libelle,
                        "montant": debit_info,
                        "formule" : ""
                    })
                if compte_credit:
                    current_entry["comptes"].append({
                        "numero": str(compte_credit),
                        "sens": "credit",
                        "intitule": libelle,
                        "montant": credit_info,
                        "formule" : ""
                    })

        if current_entry and current_entry["comptes"]:
            entries.append(current_entry)

        if entries:
            result[sheet] = entries

    return result

if __name__ == "__main__":
    file_path = r"C:\Users\fnana\Downloads\20250124_IYU_Schémas_Comptables_draft_V01.17.xlsx"
    output_file = r"C:\Users\fnana\Downloads\sortie0.json"

    try:
        data = process_excel_to_json(file_path)
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        print("Conversion terminée avec succès!")
    except Exception as e:
        print(f"Une erreur est survenue: {str(e)}")