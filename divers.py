import psycopg2
import pandas as pd
from datetime import datetime
from pathlib import Path
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment

class TransactionExporter:
    def __init__(self):
        self.db_params = {
            'dbname': 'db_hswa',
            'user': 'postgres',
            'password': 'trinita',
            'host': 'localhost',
            'port': '5432'
        }
        self.conn = None
        self.cur = None

    def connect(self):
        """Établit la connexion à la base de données"""
        try:
            self.conn = psycopg2.connect(**self.db_params)
            self.cur = self.conn.cursor()
            print("Connexion à la base de données établie.")
        except psycopg2.Error as e:
            print(f"Erreur de connexion à la base de données: {e}")
            raise

    def disconnect(self):
        """Ferme la connexion à la base de données"""
        if self.cur:
            self.cur.close()
        if self.conn:
            self.conn.close()
            print("Connexion à la base de données fermée.")

    def execute_query(self, query):
        """Exécute une requête et retourne les résultats"""
        try:
            self.cur.execute(query)
            return self.cur.fetchall()
        except psycopg2.Error as e:
            print(f"Erreur lors de l'exécution de la requête: {e}")
            raise

    def get_eug_total(self):
        """Récupère le total des opérations EUG"""
        query = """
        SELECT COALESCE(SUM(montant_a_percevoir), 0)
        FROM v_transactions_eug
        """
        result = self.execute_query(query)
        return result[0][0]

    def get_emipho_totals(self):
        """Récupère les totaux des opérations EMIPHO"""
        envois_query = """
        SELECT COALESCE(SUM(montant_total), 0)
        FROM v_transactions_emipho
        WHERE type_operation = 'ENVOI'
        """
        paiements_query = """
        SELECT COALESCE(SUM(montant_a_percevoir), 0)
        FROM v_transactions_emipho
        WHERE type_operation = 'PAIEMENT'
        """
        envois = self.execute_query(envois_query)[0][0]
        paiements = self.execute_query(paiements_query)[0][0]
        return envois, paiements

    def get_netx_totals(self):
        """Récupère les totaux des opérations NET XPRESS par pays"""
        envois_query = """
        SELECT
            pays,
            COALESCE(SUM(montant_total), 0) as total_envois
        FROM v_transactions_netx
        WHERE type_operation = 'ENVOI'
        GROUP BY pays
        """
        paiements_query = """
        SELECT
            pays,
            COALESCE(SUM(montant_a_percevoir), 0) as total_paiements
        FROM v_transactions_netx
        WHERE type_operation = 'PAIEMENT'
        GROUP BY pays
        """
        envois = pd.DataFrame(self.execute_query(envois_query), columns=['pays', 'total_envois'])
        paiements = pd.DataFrame(self.execute_query(paiements_query), columns=['pays', 'total_paiements'])
        return envois, paiements

    def get_thunes_totals(self):
        """Récupère les totaux des opérations THUNES"""
        valid_query = """
        SELECT
            devise_source,
            COUNT(*) as nombre_transactions,
            COALESCE(SUM(montant), 0) as total_montant,
            COALESCE(SUM(commission_partenaire), 0) as total_commission,
            COALESCE(SUM(montant_total), 0) as total_global
        FROM v_transactions_thunes
        WHERE statut != 'Annulé'
        GROUP BY devise_source
        """
        cancelled_query = """
        SELECT
            devise_source,
            COUNT(*) as nombre_annulations,
            COALESCE(SUM(montant), 0) as total_montant_annule,
            COALESCE(SUM(commission_partenaire), 0) as total_commission_annulee,
            COALESCE(SUM(montant_total), 0) as total_global_annule
        FROM v_transactions_thunes
        WHERE statut = 'Annulé'
        GROUP BY devise_source
        """
        valid = pd.DataFrame(self.execute_query(valid_query),
                             columns=['devise', 'nb_transactions', 'montant', 'commission', 'total'])
        cancelled = pd.DataFrame(self.execute_query(cancelled_query),
                                 columns=['devise', 'nb_annulations', 'montant_annule', 'commission_annulee', 'total_annule'])
        return valid, cancelled

    def export_to_excel(self):
        """Exporte toutes les données vers un fichier Excel avec style"""
        try:
            output_dir = Path.home() / 'documents' / 'etats_bruts' / 'exports'
            output_dir.mkdir(parents=True, exist_ok=True)
            filename = f"RECAP_JOURNEE_{datetime.now().strftime('%d_%m_%Y')}.xlsx"
            filepath = output_dir / filename

            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                eug_total = self.get_eug_total()
                emipho_envois, emipho_paiements = self.get_emipho_totals()
                netx_envois, netx_paiements = self.get_netx_totals()
                thunes_valid, thunes_cancelled = self.get_thunes_totals()

                recap_data = {
                    'Opération': [
                        'EUG - Total',
                        'EMIPHO - Envois',
                        'EMIPHO - Paiements',
                        'THUNES - Transactions Valides',
                        'THUNES - Transactions Annulées'
                    ],
                    'Montant': [
                        eug_total,
                        emipho_envois,
                        emipho_paiements,
                        thunes_valid['total'].sum(),
                        thunes_cancelled['total_annule'].sum()
                    ]
                }

                recap_df = pd.DataFrame(recap_data)
                recap_df.to_excel(writer, sheet_name='RECAP', index=False)

                for pays in netx_envois['pays'].unique():
                    envois_pays = netx_envois[netx_envois['pays'] == pays]
                    paiements_pays = netx_paiements[netx_paiements['pays'] == pays]
                    sheet_data = envois_pays.merge(paiements_pays, on='pays', how='outer')
                    sheet_data.to_excel(writer, sheet_name=f'NETX_{pays}', index=False)

                thunes_valid.to_excel(writer, sheet_name='THUNES Valides', index=False)
                thunes_cancelled.to_excel(writer, sheet_name='THUNES Annulées', index=False)

                workbook = writer.book
                for sheet_name in writer.sheets:
                    worksheet = writer.sheets[sheet_name]
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    header_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    header_alignment = Alignment(horizontal='center', vertical='center')

                    for cell in worksheet[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.border = header_border
                        cell.alignment = header_alignment

                    for column_cells in worksheet.columns:
                        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
                        worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2

                    border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                        for cell in row:
                            cell.border = border

            print(f"Export terminé. Fichier créé: {filepath}")
            return filepath
        except Exception as e:
            print(f"Erreur lors de l'export Excel: {e}")
            raise

    def run(self):
        """Exécute tout le processus d'export"""
        try:
            self.connect()
            return self.export_to_excel()
        except Exception as e:
            print(f"Une erreur est survenue: {e}")
            raise
        finally:
            self.disconnect()

if __name__ == "__main__":
    exporter = TransactionExporter()
    try:
        filepath = exporter.run()
        print(f"Fichier exporté avec succès: {filepath}")
    except Exception as e:
        print(f"Erreur lors de l'export: {e}")